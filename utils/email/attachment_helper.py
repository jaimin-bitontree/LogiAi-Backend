import fitz  # PyMuPDF
import io
import re
import os
import openpyxl
import pytesseract
import logging
from PIL import Image

# Initialize logger
logger = logging.getLogger(__name__)

# Only set Windows path if running locally on Windows
# On Render (Linux), tesseract is installed via apt and found automatically
if os.name == 'nt':
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'


def _is_meaningful_text(text: str) -> bool:
    """
    Returns True if text contains real English words.
    Returns False if text is garbage characters from image-based PDF.
    """
    real_words = re.findall(r'[a-zA-Z]{3,}', text)
    is_meaningful = len(real_words) >= 5

    if not is_meaningful and text:
        logger.debug(f"Text not meaningful: {len(real_words)} real words found in {len(text)} chars")

    return is_meaningful


# OCR
def _extract_text_from_scanned_pdf(pdf_bytes: bytes) -> str:
    """
    OCR fallback for scanned / image-based PDFs with layout preservation.
    Uses fitz to convert pages to images, then pytesseract reads them.
    """
    full_text = ""

    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for i, page in enumerate(doc):
                try:
                    # Convert page to image using fitz at 300 DPI
                    pix = page.get_pixmap(dpi=300)

                    # Convert fitz pixmap to PIL Image for pytesseract
                    img_bytes = pix.tobytes("png")
                    image = Image.open(io.BytesIO(img_bytes))

                    # Convert RGBA to RGB if needed (pytesseract needs RGB)
                    if image.mode == "RGBA":
                        image = image.convert("RGB")

                    # pytesseract reads text with layout preservation
                    # psm 6 = assume uniform block of text (good for documents)
                    custom_config = r'--oem 3 --psm 6'
                    page_text = pytesseract.image_to_string(image, config=custom_config).strip()

                    if page_text:
                        full_text += f"\n[Page {i + 1}]\n{page_text}\n"
                        logger.info(f"Page {i+1}: Extracted {len(page_text)} chars with OCR")
                    else:
                        logger.warning(f"Page {i+1}: pytesseract returned empty text")

                except Exception as e:
                    logger.error(f"OCR failed for page {i + 1}: {e}")
                    continue

    except Exception as e:
        logger.error(f"Could not open PDF for OCR: {e}")
        return ""

    return full_text.strip()


# PDF
def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """
    Extracts all text from PDF bytes.

    Step 1: Try fitz (PyMuPDF) — works for digital PDFs
    Step 2: If text is not meaningful → PDF is scanned/image → try pytesseract
    Step 3: If OCR also fails → return empty string → parse_node adds warning
    """
    # Step 1 — try fitz first
    text = ""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            if doc.is_encrypted:
                logger.warning("PDF is encrypted, cannot extract text")
                return "[LOCKED PDF: cannot extract text]"

            logger.debug(f"Processing PDF with {len(doc)} pages")
            for page_num, page in enumerate(doc):
                blocks = page.get_text("blocks")
                if blocks:
                    blocks.sort(key=lambda b: (b[1], b[0]))
                    for block in blocks:
                        if block[4].strip():
                            text += block[4].strip() + "\n"
                else:
                    text += page.get_text("text") + "\n"

    except Exception as e:
        logger.error(f"PDF extraction failed: {e}")
        return f"[ERROR: PDF extraction failed — {e}]"

    text = text.strip()

    # Step 2 — not meaningful means garbage/scanned → try pytesseract
    if not _is_meaningful_text(text):
        logger.warning(f"fitz text not meaningful ({len(text)} chars) — trying OCR...")
        text = _extract_text_from_scanned_pdf(pdf_bytes)

        if text:
            logger.info(f"pytesseract extracted {len(text)} chars")
        else:
            logger.warning("pytesseract also returned empty")

    return text


# Excel
def extract_text_from_excel(excel_bytes: bytes) -> str:
    """
    Extracts all text from Excel bytes using openpyxl.
    Reads every sheet, every row, every cell.
    Returns readable text representation of all sheets.
    """
    text = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        logger.debug(f"Processing Excel workbook with {len(wb.sheetnames)} sheets")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n[Sheet: {sheet_name}]\n"
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]
                if values:
                    text += " | ".join(values) + "\n"
                    row_count += 1

            logger.debug(f"Sheet '{sheet_name}': extracted {row_count} rows")

    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return f"[ERROR: Excel extraction failed — {e}]"

    logger.info(f"Excel extraction completed: {len(text)} chars")
    return text.strip()