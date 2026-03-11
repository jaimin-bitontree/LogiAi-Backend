import fitz  # PyMuPDF
import io
import re
import openpyxl
import easyocr
import numpy as np

# Initialize EasyOCR reader once — loads model on first run
_reader = easyocr.Reader(["en"], gpu=False)


def _is_meaningful_text(text: str) -> bool:
    """
    Returns True if text contains real English words.
    Returns False if text is garbage characters from image-based PDF.
    """
    real_words = re.findall(r'[a-zA-Z]{3,}', text)
    return len(real_words) >= 5



# OCR
def _extract_text_from_scanned_pdf(pdf_bytes: bytes) -> str:
    """
    OCR fallback for scanned / image-based PDFs with layout preservation.
    Uses fitz to convert pages to images, then EasyOCR reads them with coordinates.
    """
    full_text = ""

    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            for i, page in enumerate(doc):
                try:
                    # Convert page to image using fitz
                    pix = page.get_pixmap(dpi=300)

                    # Convert to numpy array for EasyOCR
                    image_np = np.frombuffer(pix.samples, dtype=np.uint8).reshape(
                        pix.height, pix.width, pix.n
                    )

                    # EasyOCR reads text with coordinates for layout preservation
                    results = _reader.readtext(
                        image_np,
                        detail=1,
                        paragraph=False,
                        width_ths=0.7,
                        height_ths=0.5,
                        ycenter_ths=0.5,
                        decoder='beamsearch',
                        batch_size=1,
                        contrast_ths=0.3,
                        adjust_contrast=0.7,
                        text_threshold=0.6,
                        low_text=0.3,
                        link_threshold=0.3,
                        mag_ratio=1.0
                    )

                    if not results:
                        continue

                    # Sort by reading order (top to bottom, left to right)
                    results.sort(key=lambda x: (x[0][0][1], x[0][0][0]))

                    # Group text by lines based on y-coordinate
                    lines = []
                    current_line = []
                    current_y = None
                    y_threshold = 20

                    for bbox, text, confidence in results:
                        top_y = bbox[0][1]

                        if current_y is None:
                            current_y = top_y
                            current_line.append(text)
                        elif abs(top_y - current_y) <= y_threshold:
                            current_line.append(text)
                        else:
                            if current_line:
                                lines.append(" ".join(current_line))
                            current_line = [text]
                            current_y = top_y

                    if current_line:
                        lines.append(" ".join(current_line))

                    page_text = "\n".join(lines).strip()

                    if page_text:
                        full_text += f"\n[Page {i + 1}]\n{page_text}\n"
                        print(f"[attachment_helper] ✅ Page {i+1}: Extracted {len(page_text)} chars")

                except Exception as e:
                    print(f"[attachment_helper] ❌ OCR failed for page {i + 1} — {e}")
                    continue

    except Exception as e:
        print(f"[attachment_helper] ❌ Could not open PDF for OCR — {e}")
        return ""

    return full_text.strip()


# PDF
def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """
    Extracts all text from PDF bytes.

    Step 1: Try fitz (PyMuPDF) — works for digital PDFs
    Step 2: If text is not meaningful → PDF is scanned/image → try EasyOCR
    Step 3: If OCR also fails → return empty string → parse_node adds warning
    """
    # Step 1 — try fitz first
    text = ""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            if doc.is_encrypted:
                return "[LOCKED PDF: cannot extract text]"

            for page_num, page in enumerate(doc):
                blocks = page.get_text("blocks")
                if blocks:
                    blocks.sort(key=lambda b: (b[1], b[0]))
                    for block in blocks:
                        if block[4].strip():
                            text += block[4].strip() + "\n"
                else:
                    text += page.get_text("text") + "\n"

    except Exception as e:
        return f"[ERROR: PDF extraction failed — {e}]"

    text = text.strip()

    # Step 2 — not meaningful means garbage/scanned → try EasyOCR
    if not _is_meaningful_text(text):
        print(f"[attachment_helper] ⚠️ fitz text not meaningful ({len(text)} chars) — trying EasyOCR...")
        text = _extract_text_from_scanned_pdf(pdf_bytes)

        if text:
            print(f"[attachment_helper] ✅ EasyOCR extracted {len(text)} chars")
        else:
            print("[attachment_helper] ❌ EasyOCR also returned empty")

    return text


# Excel
def extract_text_from_excel(excel_bytes: bytes) -> str:
    """
    Extracts all text from Excel bytes using openpyxl.
    Reads every sheet, every row, every cell.
    Returns readable text representation of all sheets.
    """
    text = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text += f"\n[Sheet: {sheet_name}]\n"
            for row in ws.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]
                if values:
                    text += " | ".join(values) + "\n"
    except Exception as e:
        return f"[ERROR: Excel extraction failed — {e}]"
    return text.strip()