from email import policy
from email.parser import BytesParser
from models.shipment import Attachment
from utils.email_utils import (
    extract_email_address,
    extract_body,
    clean_email_body,
    extract_attachments,
)
from agent.state import AgentState
import fitz
import io
import openpyxl
from langchain_core.tools import tool


# pdf

def _extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extracts all text from PDF bytes using PyMuPDF (fitz)."""
    text = ""
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            text += page.get_text() + "\n"
    return text.strip()


# excel

def _extract_text_from_excel(excel_bytes: bytes) -> str:
    """
    Extracts all text from Excel bytes using openpyxl.
    Reads every sheet, every row, every cell.
    Returns readable text representation of all sheets.
    """
    text = ""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text += f"\n[Sheet: {sheet_name}]\n"

        for row in ws.iter_rows(values_only=True):
            # filter out completely empty rows
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip() != ""]
            if values:
                text += " | ".join(values) + "\n"

    return text.strip()


# tool — append PDF text to body

@tool
def append_pdf_text_to_body(body: str, pdf_filename: str, pdf_text: str) -> str:
    """
    Appends extracted PDF text to the email body.
    Called once per PDF attachment found in the email.
    Returns the combined body string.
    """
    if pdf_text:
        return body + f"\n\n[PDF: {pdf_filename}]\n{pdf_text}"
    return body


# tool — append Excel text to body

@tool
def append_excel_text_to_body(body: str, excel_filename: str, excel_text: str) -> str:
    """
    Appends extracted Excel text to the email body.
    Called once per Excel attachment found in the email.
    Returns the combined body string.
    """
    if excel_text:
        return body + f"\n\n[EXCEL: {excel_filename}]\n{excel_text}"
    return body


# langgraph node

def parser_node(state: AgentState) -> AgentState:

    raw_email = state["raw_email"]
    msg = BytesParser(policy=policy.default).parsebytes(raw_email)

    # Subject
    subject = msg.get("Subject", "")

    # Sender
    sender_raw = msg.get("From", "")
    customer_email = extract_email_address(sender_raw)

    # IDs
    message_id = (msg.get("Message-ID", "") or "").strip().strip("<>")
    in_reply_to = msg.get("In-Reply-To", "")
    thread_id = in_reply_to.strip().strip("<>") if in_reply_to else None

    # Body
    raw_body = extract_body(msg)
    clean_body = clean_email_body(raw_body)

    # Attachments — convert raw dicts to Attachment objects
    raw_attachments = extract_attachments(msg)
    attachments = [
        Attachment(
            filename=a["filename"],
            content_type=a["content_type"],
            content=a.get("content")  # keep the file bytes
        )
        for a in raw_attachments
    ]

    # PDF + Excel extraction — append to body via @tool
    updated_body = clean_body
    for att in attachments:
        filename = (att.filename or "").lower()

        is_pdf = filename.endswith(".pdf") or att.content_type == "application/pdf"
        is_excel = (
            filename.endswith(".xlsx") or
            filename.endswith(".xls") or
            att.content_type in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            )
        )

        if is_pdf and att.content:
            try:
                print(f"[parse_node] Extracting PDF: {att.filename}")
                pdf_text = _extract_text_from_pdf(att.content)
                if not pdf_text:
                    print(f"[parse_node] ⚠️ No text extracted from {att.filename} — may be scanned or image-based PDF")
                    updated_body += f"\n\n[WARNING: Could not read {att.filename} — please resend as a digital PDF]"
                else:
                    updated_body = append_pdf_text_to_body.invoke({
                        "body": updated_body,
                        "pdf_filename": att.filename,
                        "pdf_text": pdf_text,
                    })
                    print(f"[parse_node] ✅ Extracted {len(pdf_text)} chars from {att.filename}")
            except Exception as e:
                print(f"[parse_node] ❌ PDF extraction failed: {att.filename} — {e}")
                updated_body += f"\n\n[WARNING: Could not read {att.filename} — please resend as a digital PDF]"

        elif is_excel and att.content:
            try:
                print(f"[parse_node] Extracting Excel: {att.filename}")
                excel_text = _extract_text_from_excel(att.content)
                if not excel_text:
                    print(f"[parse_node] ⚠️ No text extracted from {att.filename} — file may be empty")
                    updated_body += f"\n\n[WARNING: Could not read {att.filename} — please resend as a valid Excel file]"
                else:
                    updated_body = append_excel_text_to_body.invoke({
                        "body": updated_body,
                        "excel_filename": att.filename,
                        "excel_text": excel_text,
                    })
                    print(f"[parse_node] ✅ Extracted {len(excel_text)} chars from {att.filename}")
            except Exception as e:
                print(f"[parse_node] ❌ Excel extraction failed: {att.filename} — {e}")
                updated_body += f"\n\n[WARNING: Could not read {att.filename} — please resend as a valid Excel file]"

    message_ids = state.get("message_ids")
    if not isinstance(message_ids, list):
        message_ids = []

    if message_id and message_id not in message_ids:
        message_ids.append(message_id)
        print(message_ids)

    # Update state
    state.update({
        "message_ids": message_ids,
        "last_message_id": message_id,
        "thread_id": message_id,
        "conversation_id": thread_id,
        "customer_email": customer_email,
        "subject": subject,
        "body": updated_body,
        "attachments": attachments,
    })

    return state
